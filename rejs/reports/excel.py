from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import date


class ExcelExporter:
	def __init__(self, filename):
		self.wb = Workbook()
		self.filename = filename

	def save(self):
		self.wb.save(self.filename)

	# ---------- ZAŁOGA ----------
	def add_zaloga(self, rows):
		ws = self.wb.active
		ws.title = "Załoga"

		if not rows:
			return

		headers = rows[0].keys()
		ws.append(list(headers))

		for cell in ws[1]:
			cell.font = Font(bold=True)

		for r in rows:
			ws.append(list(r.values()))

	# ---------- WACHTY ----------
	def add_wachty(self, wachty):
		ws = self.wb.create_sheet("Wachty")

		for w in wachty:
			ws.append([f"Wachta: {w['nazwa']}"])
			ws.append(["Imię", "Nazwisko", "Rola"])

			for c in w["czlonkowie"]:
				ws.append([c["imie"], c["nazwisko"], c["rola"]])

			ws.append([])

	# ---------- WPŁATY ----------
	def add_wplaty(self, rows):
		ws = self.wb.create_sheet("Wpłaty")

		if not rows:
			return

		headers = rows[0].keys()
		ws.append(list(headers))

		for cell in ws[1]:
			cell.font = Font(bold=True)

		for r in rows:
			ws.append(list(r.values()))

	# ---------- DANE WRAŻLIWE ----------
	def add_dane_wrazliwe(self, rows):
		if not rows:
			return

		ws = self.wb.create_sheet("Dane wrażliwe")

		headers = rows[0].keys()
		ws.append(list(headers))

		for cell in ws[1]:
			cell.font = Font(bold=True)

		for r in rows:
			ws.append(list(r.values()))

	# ---------- CREW LIST (IMO FAL 5) ----------
	def add_crew_list(self, rows):
		ws = self.wb.create_sheet("Crew List")

		thin = Side(style="thin")
		border = Border(left=thin, right=thin, top=thin, bottom=thin)

		def style_cell(cell, bold=False, center=False, wrap=False):
			cell.border = border
			cell.font = Font(bold=bold)
			cell.alignment = Alignment(
				horizontal="center" if center else "left",
				vertical="center",
				wrap_text=wrap
			)

		# --- SZEROKOŚCI KOLUMN (zbliżone do IMO FAL 5) ---
		widths = {
			"A": 4,
			"B": 18,
			"C": 18,
			"D": 6,
			"E": 14,
			"F": 20,
			"G": 16,
			"H": 20,
			"I": 18,
			"J": 20,
			"K": 6,
		}
		for col, w in widths.items():
			ws.column_dimensions[col].width = w

		# --- TYTUŁ ---
		ws.merge_cells("E1:G1")
		ws["E1"] = "CREW LIST"
		ws["E1"].font = Font(bold=True)
		ws["E1"].alignment = Alignment(horizontal="center")

		ws["K2"] = "Page No."
		ws["K3"] = 1

		# --- NAGŁÓWKI TABELI ---
		headers = [
			"No.",
			"Family name",
			"Given names",
			"Age",
			"Date of birth",
			"Place of birth",
			"Nationality",
			"Capacity or rank",
			"Type of identity document",
			"Number of identity document",
			"Sex",
		]

		start_row = 8
		for col, header in enumerate(headers, start=1):
			cell = ws.cell(row=start_row, column=col)
			cell.value = header
			style_cell(cell, bold=True, center=True, wrap=True)

		# --- DANE ---
		data_row = start_row + 1

		for idx, r in enumerate(rows, start=1):
			dob = r.get("date_of_birth")

			values = [
				idx,
				r.get("family_name", ""),
				r.get("given_names", ""),
				r.get("age", ""),
				dob.strftime("%d.%m.%Y") if dob else "",
				r.get("place_of_birth", ""),
				r.get("nationality", ""),
				r.get("rank", ""),
				r.get("document_type", ""),
				r.get("document_number", ""),
				r.get("sex", ""),
			]

			for col, val in enumerate(values, start=1):
				cell = ws.cell(row=data_row, column=col)
				cell.value = val
				style_cell(cell, center=(col in [1, 4, 11]))

			data_row += 1

		# --- RAMKI ---
		for row in ws.iter_rows(
			min_row=start_row,
			max_row=data_row - 1,
			min_col=1,
			max_col=11
		):
			for cell in row:
				cell.border = border
